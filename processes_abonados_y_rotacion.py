import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from gcp_paths import path_abonados, path_rotacion, path_transparencia
from gcp_utils import read, upload_files_to_gcs, parkings_current_info


def main():
    events = []
    registros = []

    for parking_name, [parking_id, _, _, parking_norm] in parkings_current_info.items():
        df_a, h_a = get_data(parking_norm, 1, clean_abonados)
        df_r, h_r = get_data(parking_norm, 2, clean_rotacion)

        registros.append(build_records(df_a, parking_id, parking_name, 'ABONADO'))
        registros.append(build_records(df_r, parking_id, parking_name, 'ROTACIÓN'))

        events += [
            [parking_id, {'Horas Abonados': h_a}],
            [parking_id, {'Horas Rotación': h_r}],
        ]

    all_rec = pd.concat(registros, ignore_index=True)
    create_transparencia(all_rec, path_transparencia)
    
    return events


def get_data(parking, tipo, clean_func):
    ss = path_abonados if tipo == 1 else path_rotacion
    ss = ss.replace('&', parking)
    df = read(ss, 'excel')
    clean = clean_func(df)
    horas = clean['tiempo'].sum()
    return clean, horas


def clean_rotacion(df):
    df = df.loc[df['MovementType'] == 1].copy()
    df['Time'] = pd.to_datetime(df['Time'])
    df['PaidFrom'] = pd.to_datetime(df['PaidFrom'])
    df = df.sort_values('Time')
    df['tiempo'] = (df['Time'] - df['PaidFrom']).dt.total_seconds() / 3600
    return df[['Time', 'PaidFrom', 'tiempo']]


def clean_abonados(df):
    df = df.sort_values(['UserNo', 'Time']).reset_index(drop=True)
    df['Time'] = pd.to_datetime(df['Time'])

    prev_mt   = df['MovementType'].shift(1)
    prev_usr  = df['UserNo'].shift(1)
    # Filtrar salidas huérfanas y transacciones
    mask_keep = ~((df['MovementType'] == 4) & (df['UserNo'] != prev_usr))
    df = df.loc[mask_keep & (df['MovementType'] != 1)].reset_index(drop=True)

    prev_mt   = df['MovementType'].shift(1)
    prev_usr  = df['UserNo'].shift(1)
    # Entradas válidas
    valid     = (df['MovementType'] == 4) & (prev_mt == 0) & (df['UserNo'] == prev_usr)
    df['Entrada'] = pd.NaT
    df.loc[valid, 'Entrada'] = df['Time'].shift(1)

    df = df.dropna(subset=['Entrada']).reset_index(drop=True)
    df['tiempo'] = (df['Time'] - df['Entrada']).dt.total_seconds() / 3600
    return df[['Time', 'Entrada', 'tiempo']]


def build_records(df, gid, parking, tipo):
    lbl = 'SAN BERNARDO' if parking == 'SB' else parking.upper()
    ent = pd.DataFrame({
        'Time': df.iloc[:, 1],             # Entrada o PaidFrom
        'MovementTypeDesig': 'ENTRADA',
        'TIPO': tipo
    })
    sal = pd.DataFrame({
        'Time': df['Time'],
        'MovementTypeDesig': 'SALIDA',
        'TIPO': tipo
    })
    out = pd.concat([ent, sal], ignore_index=True)
    out['GarageNo'] = gid
    out['APARCAMIENTO'] = lbl
    return out


def create_transparencia(records, output_path_gcs):
    local_path = '/tmp/temp_transparencia.xlsx'
    web_path = '/tmp/temp_transparencia_WEB.xlsx'

    # Mes de estudio
    m = records['Time'].dt.month.value_counts().idxmax()
    temp = records.loc[records['Time'].dt.month.eq(m)].sort_values('Time')
    temp.to_excel(local_path, index=False)

    # --- aquí sólo un load_workbook y un solo save para ambas versiones ---
    wb     = load_workbook(local_path, data_only=True)
    sheet  = wb.active

    # Predefinir estilos
    hdr_f = Font(bold=True, color="002060", size=11, name="Rubik")
    hdr_F = PatternFill(start_color="ddebf7", end_color="ddebf7", fill_type="solid")
    spc_f = Font(bold=True, color="ffff00", size=11, name="Rubik")
    spc_F = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
    bd_f  = Font(size=11, name="Rubik")
    al    = Alignment(horizontal="center", vertical="center")

    # Índice de columna Time
    time_idx = next(c.col_idx for c in sheet[1] if c.value == 'Time')

    # Pase único de estilizado y formato fecha
    for r in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                             min_col=1, max_col=sheet.max_column):
        for cell in r:
            if cell.row == 1:
                if cell.value == 'APARCAMIENTO':
                    cell.font, cell.fill = spc_f, spc_F
                else:
                    cell.font, cell.fill = hdr_f, hdr_F
            else:
                cell.font = bd_f
                if cell.col_idx == time_idx:
                    cell.number_format = 'DD/MM/YYYY HH:MM'
            cell.alignment = al

    # Anchos de columnas
    for col, w in (('A',20),('B',25),('C',25),('D',20),('E',20)): sheet.column_dimensions[col].width = w

    # Guardar versión normal
    wb.save(local_path)
    # Borrar cols 3 y 4, ajustar anchos y guardar versión WEB
    sheet.delete_cols(3, 2)
    for col, w in (('A',20),('B',25),('C',20)):
        sheet.column_dimensions[col].width = w
    wb.save(web_path)

    # Subir ambos archivos a GCS
    upload_files_to_gcs({
        output_path_gcs: local_path,
        output_path_gcs.replace('.xlsx', '_WEB.xlsx'): web_path
    })
